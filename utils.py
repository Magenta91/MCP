import os
import json
import pandas as pd
from typing import Dict, List, Any, Optional
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload, MediaIoBaseUpload
from io import BytesIO, StringIO
import openpyxl
from openpyxl import Workbook


def get_drive_service():
    """Initialize Google Drive service with service account credentials."""
    key_path = os.getenv('GOOGLE_SERVICE_ACCOUNT_KEY_PATH')
    if not key_path or not os.path.exists(key_path):
        raise ValueError("Service account key file not found")
    
    credentials = service_account.Credentials.from_service_account_file(
        key_path,
        scopes=['https://www.googleapis.com/auth/drive']
    )
    return build('drive', 'v3', credentials=credentials)


def download_file_from_drive(file_id: str, service) -> bytes:
    """Download file content from Google Drive."""
    try:
        request = service.files().get_media(fileId=file_id)
        file_content = request.execute()
        return file_content
    except Exception as e:
        raise Exception(f"Failed to download file: {str(e)}")


def upload_to_drive(content: bytes, filename: str, service, folder_id: str) -> str:
    """Upload content to Google Drive folder."""
    try:
        # Determine MIME type based on file extension
        if filename.endswith('.json'):
            mimetype = 'application/json'
        elif filename.endswith('.xlsx'):
            mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        elif filename.endswith('.html'):
            mimetype = 'text/html'
        else:
            mimetype = 'application/octet-stream'
            
        media = MediaIoBaseUpload(BytesIO(content), mimetype=mimetype)
        file_metadata = {
            'name': filename,
            'parents': [folder_id]
        }
        
        file = service.files().create(
            body=file_metadata,
            media_body=media,
            fields='id'
        ).execute()
        
        return file.get('id')
    except Exception as e:
        raise Exception(f"Failed to upload file: {str(e)}")


def upload_to_drive_file(file_path: str, service, folder_id: str) -> str:
    """Upload file from local path to Google Drive folder."""
    try:
        filename = os.path.basename(file_path)
        
        # Read file content and upload as bytes
        with open(file_path, 'rb') as f:
            file_content = f.read()
        
        # Determine MIME type based on file extension
        if filename.endswith('.xlsx'):
            mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        elif filename.endswith('.json'):
            mimetype = 'application/json'
        else:
            mimetype = 'application/octet-stream'
        
        media = MediaIoBaseUpload(BytesIO(file_content), mimetype=mimetype)
        file_metadata = {
            'name': filename,
            'parents': [folder_id]
        }
        
        file = service.files().create(
            body=file_metadata,
            media_body=media,
            fields='id'
        ).execute()
        
        return file.get('id')
    except Exception as e:
        raise Exception(f"Failed to upload file: {str(e)}")


def list_files_in_folder(service, folder_id: str) -> List[Dict]:
    """List all files in a Google Drive folder."""
    try:
        results = service.files().list(
            q=f"'{folder_id}' in parents and trashed=false",
            fields="files(id, name, mimeType, createdTime, modifiedTime, size)"
        ).execute()
        
        return results.get('files', [])
    except Exception as e:
        raise Exception(f"Failed to list files: {str(e)}")


def extract_metadata_from_dataframe(df: pd.DataFrame, filename: str) -> Dict[str, Any]:
    """Extract metadata from pandas DataFrame."""
    metadata = {
        "filename": filename,
        "row_count": len(df),
        "column_count": len(df.columns),
        "columns": [],
        "summary_stats": {}
    }
    
    for col in df.columns:
        col_info = {
            "name": col,
            "data_type": str(df[col].dtype),
            "null_count": int(df[col].isnull().sum()),
            "null_percentage": float(df[col].isnull().sum() / len(df) * 100),
            "unique_count": int(df[col].nunique())
        }
        
        # Add basic statistics for numeric columns
        if df[col].dtype in ['int64', 'float64', 'int32', 'float32']:
            col_info.update({
                "min_value": float(df[col].min()) if not df[col].empty else None,
                "max_value": float(df[col].max()) if not df[col].empty else None,
                "mean_value": float(df[col].mean()) if not df[col].empty else None,
                "std_value": float(df[col].std()) if not df[col].empty else None
            })
        
        metadata["columns"].append(col_info)
    
    return metadata


def suggest_dq_rules(metadata: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Suggest data quality rules based on metadata."""
    dq_rules = []
    
    for col in metadata["columns"]:
        col_name = col["name"]
        
        # Not null rule for columns with low null percentage
        if col["null_percentage"] < 10:
            dq_rules.append({
                "column": col_name,
                "rule_type": "not_null",
                "description": f"Column '{col_name}' should not contain null values",
                "severity": "error"
            })
        
        # Uniqueness rule for columns with high unique count
        if col["unique_count"] / metadata["row_count"] > 0.95:
            dq_rules.append({
                "column": col_name,
                "rule_type": "unique",
                "description": f"Column '{col_name}' should contain unique values",
                "severity": "warning"
            })
        
        # Range validation for numeric columns
        if "min_value" in col and "max_value" in col:
            dq_rules.append({
                "column": col_name,
                "rule_type": "range",
                "description": f"Column '{col_name}' values should be between {col['min_value']} and {col['max_value']}",
                "min_value": col["min_value"],
                "max_value": col["max_value"],
                "severity": "warning"
            })
    
    return dq_rules


def create_contract_excel(metadata: Dict[str, Any], dq_rules: List[Dict[str, Any]], output_path: str):
    """Create contract Excel file with schema and DQ information."""
    wb = Workbook()
    
    # Schema sheet
    ws_schema = wb.active
    ws_schema.title = "Schema"
    ws_schema.append(["Column Name", "Data Type", "Null Count", "Null %", "Unique Count", "Min Value", "Max Value", "Mean", "Std Dev"])
    
    for col in metadata["columns"]:
        row = [
            col["name"],
            col["data_type"],
            col["null_count"],
            f"{col['null_percentage']:.2f}%",
            col["unique_count"],
            col.get("min_value", ""),
            col.get("max_value", ""),
            col.get("mean_value", ""),
            col.get("std_value", "")
        ]
        ws_schema.append(row)
    
    # Data Quality Rules sheet
    ws_dq = wb.create_sheet("Data_Quality_Rules")
    ws_dq.append(["Column", "Rule Type", "Description", "Severity", "Min Value", "Max Value"])
    
    for rule in dq_rules:
        row = [
            rule["column"],
            rule["rule_type"],
            rule["description"],
            rule["severity"],
            rule.get("min_value", ""),
            rule.get("max_value", "")
        ]
        ws_dq.append(row)
    
    # Summary sheet
    ws_summary = wb.create_sheet("Summary")
    ws_summary.append(["Metric", "Value"])
    ws_summary.append(["Filename", metadata["filename"]])
    ws_summary.append(["Total Rows", metadata["row_count"]])
    ws_summary.append(["Total Columns", metadata["column_count"]])
    ws_summary.append(["DQ Rules Count", len(dq_rules)])
    
    wb.save(output_path)


def generate_dq_report(metadata: Dict[str, Any], dq_rules: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Generate data quality report."""
    report = {
        "dataset_info": {
            "filename": metadata["filename"],
            "row_count": metadata["row_count"],
            "column_count": metadata["column_count"]
        },
        "data_quality_summary": {
            "total_rules": len(dq_rules),
            "error_rules": len([r for r in dq_rules if r["severity"] == "error"]),
            "warning_rules": len([r for r in dq_rules if r["severity"] == "warning"])
        },
        "column_quality": [],
        "suggested_rules": dq_rules
    }
    
    for col in metadata["columns"]:
        col_quality = {
            "column_name": col["name"],
            "completeness": 100 - col["null_percentage"],
            "uniqueness": (col["unique_count"] / metadata["row_count"]) * 100,
            "data_type": col["data_type"]
        }
        report["column_quality"].append(col_quality)
    
    return report


def publish_to_mock_catalog(metadata: Dict[str, Any], contract_path: str, 
                          dq_report: Dict[str, Any], drive_service, catalog_folder_id: str) -> Dict[str, str]:
    """Publish all artifacts to mock catalog (MCP_client folder)."""
    uploaded_files = {}
    
    try:
        # Upload metadata.json
        metadata_content = json.dumps(metadata, indent=2).encode('utf-8')
        metadata_file_id = upload_to_drive(
            metadata_content, 
            f"{metadata['filename']}_metadata.json", 
            drive_service, 
            catalog_folder_id
        )
        uploaded_files["metadata"] = metadata_file_id
        
        # Upload contract.xlsx
        contract_file_id = upload_to_drive_file(contract_path, drive_service, catalog_folder_id)
        uploaded_files["contract"] = contract_file_id
        
        # Upload dq_report.json
        dq_report_content = json.dumps(dq_report, indent=2).encode('utf-8')
        dq_report_file_id = upload_to_drive(
            dq_report_content, 
            f"{metadata['filename']}_dq_report.json", 
            drive_service, 
            catalog_folder_id
        )
        uploaded_files["dq_report"] = dq_report_file_id
        
        return uploaded_files
        
    except Exception as e:
        raise Exception(f"Failed to publish to catalog: {str(e)}")